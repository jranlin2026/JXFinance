from __future__ import annotations

import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def normalize(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if value is None:
        return ""
    return value


def main():
    if len(sys.argv) != 4:
        raise SystemExit("Usage: extractWorkbookRows.py <input.xlsx> <output.json> <fields-json-or-null>")

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    fields_arg = sys.argv[3]
    fields = None if fields_arg == "null" else json.loads(fields_arg)

    readable_path = input_path
    temp_path = None
    if input_path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        temp_path = output_path.with_suffix(".source.xlsx")
        shutil.copyfile(input_path, temp_path)
        readable_path = temp_path

    workbook = None
    try:
        workbook = load_workbook(readable_path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            output_path.write_text(json.dumps({"rows": [], "missing": []}, ensure_ascii=False), encoding="utf-8")
            return

        headers = [str(value).strip() if value is not None else "" for value in header_row]
        selected_fields = fields or headers
        indexes = []
        missing = []
        for field in selected_fields:
            if field in headers:
                indexes.append(headers.index(field))
            else:
                indexes.append(None)
                missing.append(field)

        rows = []
        for source_row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None and str(value).strip() != "" for value in source_row):
                continue
            row = {}
            for field, index in zip(selected_fields, indexes):
                row[field] = "" if index is None or index >= len(source_row) else normalize(source_row[index])
            rows.append(row)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps({"rows": rows, "missing": missing}, ensure_ascii=False), encoding="utf-8")
    finally:
        if workbook is not None:
            workbook.close()
        if temp_path:
            try:
                temp_path.unlink(missing_ok=True)
            except PermissionError:
                pass


if __name__ == "__main__":
    main()
